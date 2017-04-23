# -*- coding: ISO-8859-1 -*-
from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, logging, gui #import some libraries from PsychoPy
from psychopy import *
from psychopy.constants import *  # things like STARTED, FINISHED
import random #import some libraries from python
from random import choice
import array
import numpy
import os  # handy system and path functions
import pylab
try:
    import openpyxl
    from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl=True
except:
    haveOpenpyxl=False
    
#variables for working memory tasks
#listOfConsonants = array.array('c','BCDFGHKLMNPRSTVWZ') #list of consonants to be used during phonological and visuospatial wm trials
listOfConsonants =  ['B','C','D','F','G','H','J','K','L','M','N','P','R','S','T','Z'] #list of consonants to be used during the experi
testLetters = ['b','c','d','f','g','h','j','k','l','m','n','p','r','s','t','z'] 
allLocs = [0,1,2,3,4,5,6,7,8,9,10,11,13,14,15,16,17,18,19,20,21,22,23,24]
listSize = len(listOfConsonants)
scale = 188/2
testLetterPos = random.randint(0,listSize-1) #if the letter is from the list of consonants, this picks the position from the array of consonants
#xCoor = [-5*scale/6,-scale/2,-scale/6,5*scale/6,scale/2,scale/6,-5*scale/6,-scale/2,-scale/6,5*scale/6,scale/2,scale/6,-5*scale/6,-scale/2,-scale/6,5*scale/6,scale/2,scale/6,5*scale/6,-scale/2,-scale/6,5*scale/6,scale/2,scale/6,5*scale/6,-scale/2,-scale/6,5*scale/6,scale/2,scale/6,5*scale/6,-scale/2,-scale/6,5*scale/6,scale/2,scale/6]
#yCoor = [5*scale/6,5*scale/6,5*scale/6,5*scale/6,5*scale/6,5*scale/6,scale/2,scale/2,scale/2,scale/2,scale/2,scale/2,scale/6,scale/6,scale/6,scale/6,scale/6,scale/6,-5*scale/6,-5*scale/6,-5*scale/6,-5*scale/6,-5*scale/6,-5*scale/6,-scale/2,-scale/2,-scale/2,-scale/2,-scale/2,-scale/2,-scale/6,-scale/6,-scale/6,-scale/6,-scale/6,-scale/6]
xCoor = [-4*scale/5,-2*scale/5,0,2*scale/5,4*scale/5,-4*scale/5,-2*scale/5,0,2*scale/5,4*scale/5,-4*scale/5,-2*scale/5,0,2*scale/5,4*scale/5,-4*scale/5,-2*scale/5,0,2*scale/5,4*scale/5,-4*scale/5,-2*scale/5,0,2*scale/5,4*scale/5,-4*scale/5,-2*scale/5,0,2*scale/5,4*scale/5]
yCoor = [4*scale/5,4*scale/5,4*scale/5,4*scale/5,4*scale/5,2*scale/5,2*scale/5,2*scale/5,2*scale/5,2*scale/5,0,0,0,0,0,-2*scale/5,-2*scale/5,-2*scale/5,-2*scale/5,-2*scale/5,-4*scale/5,-4*scale/5,-4*scale/5,-4*scale/5,-4*scale/5]


#variables for visuospatial working memory tasks
gridSizeX = 5 #the number of  horizontal locations a letter can pop up --> the screen is divided into invisible grids and one is chosen randomly to show the letter
gridSizeY = 5 #the number of vertical locations a letter can pop up --> the screen is divided into invisible grids and one is chosen randomly to show the letter
stepDirV = [0,0,0,0,0,0,0,0,0,0,1,2,3,4,1,2,3,4,1,2,3,4] #the four directions for the probe location of solo visuospatial trials
stepDirVC = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,2,3,4,1,2,3,4,1,2,3,4,1,2] #the four directions for the probe location of visuospatial trials with calc
random.shuffle(stepDirV)
random.shuffle(stepDirVC)

testLetter = 'O' #the test letter for visuospatial working memory answer phase


key_resp = event.BuilderKeyResponse()  # create an object of type KeyResponse
key_resp.status = NOT_STARTED


cont_key = event.BuilderKeyResponse() # for handling the breaks in between blocks
cont_key.status = NOT_STARTED

##############################################################################################
#-------------------base functions to be used during the experiment---------------------------
##############################################################################################
def my_shuffle(array):
        random.shuffle(array)
        return array
##############################################################################################
#determine the letters to be used during the current WM trial --> input parameters: number of
#letters to be tested, length of the list of letters to be used, output--> location of letters to be used 
def wmLets(span):
    i = 0
    letsUsed = numpy.zeros(span)
    #choose the first letter to be used
    letsUsed[0] = random.randint(0,listSize-1) #pick the position from the array of consonants
    #choose the letters to be used during the current working memory trial
    for i in range(0,span):
        curLet = random.randint(0,listSize-1) #pick the position from the array of consonants
        #check if the chosen letter has already been used or not
        j = 0
        while j in range(0,i):
            #print 'BURA2'
            if curLet == letsUsed[j]:
                j = -1
                curLet = random.randint(0,listSize-1) #pick a different letter from the list
            j += 1
        letsUsed[i] = curLet
        i += 1
    return letsUsed
##############################################################################################
#determine the location for letters to be used during the current WM trial --> input parameters:
#number of letters to be tested, number of possible locations
def visualLocs(span):
    global xCoor,yCoor,allLocs
    locsUsed = numpy.zeros(span) #grid number of locations to be used
    locsX = numpy.zeros(span) #x coordinate of locations to be used
    locsY = numpy.zeros(span) #y coordinate of locations to be used
    #choose the locations to be used during the current working memory trial
    for i in range(0,span):
        curLoc = choice(allLocs)
        #check if the chosen location has already been used or not
        j = 0
        while 0<=j<i:
            #print 'BURA3'
            if curLoc == locsUsed[j]:
                curLoc = choice(allLocs)
                j = -1
            j += 1          
        locsUsed[i] = curLoc
    return locsUsed
##############################################################################################
#find the prob location
def findProb(locsUsed,dir,span):
    global xCoor,yCoor,allLocs
    curLoc= choice(locsUsed)
    indx = curLoc
    tmp = 0
    #print 'curLoc',curLoc
    while curLoc in locsUsed and tmp < 20:
        tmp = tmp+1
        #print 'BURA4'
        #print 'indx',indx,'span',span,'locsUsed',locsUsed
        curLoc=choice(locsUsed)
        indx = curLoc
        if dir == 1 and curLoc >4 and curLoc != 17:
            curLoc = curLoc - 5
        elif dir  == 2 and curLoc < 20 and curLoc != 7:
            curLoc = curLoc + 5
        elif dir  == 3 and curLoc%5 != 4 and curLoc != 11:
            curLoc  = curLoc +1
        elif dir == 4 and curLoc%5 != 0 and curLoc != 13:
            curLoc = curLoc -1
        if tmp == 20 and dir == 4:
            tmp = 0
            dir = dir-1
        elif tmp == 20 and dir == 1:
            tmp = 0 
            dir = dir +1
        elif tmp == 20:
            tmp = 0 
            dir = dir -1
    return indx,curLoc,dir
##############################################################################################
#determine the location for letters to be used during the current WM trial --> input parameters:
#number of letters to be tested, number of possible locations
def phonoLocs(span):
    #choose the first location to be used
    locs =numpy.zeros(span)
    #choose the locations to be used during the current working memory trial
    for i in range(1,int(span/2)+1):
        locs[i-1] = -i*20
    locs[int(span/2)] = 0
    for j in range(int(span/2)+1,span):
        locs[j] = (j-int(span/2))*20
    return locs
##############################################################################################
#create the stimuli to be used in the current experiment
def create_stimuli(participant,group,calcFile,type,reps,spans):
    corrAnsSide = participant%2 #left if participant no is odd, right if even
    # Setup files for saving
    if not os.path.isdir('mydata'):
        os.makedirs('mydata')  # if this fails (e.g. permissions) we will get error
    filename = 'mydata' + os.path.sep + 'subject_'+ str(participant) + '_session_' + type
    logFile = logging.LogFile(filename+'.log', level=logging.EXP)
    logging.console.setLevel(logging.WARNING)  # this outputs to the screen, not a file

    #file path
    file_name = 'C:\\Users\\user\\Desktop\\exp\\wm.xlsx'
    calcWb = load_workbook(calcFile)
    wb = load_workbook('wm.xlsx')
    sheet = wb.create_sheet(0)
    sheet.title = 'subject_'+str(participant) + '_' +type + "_stimuli"
    
    calcSheet = calcWb.get_active_sheet()
    rowNo = calcSheet.get_highest_row()
    probAnsSide = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,1,1,1,1,1,1,1,1,1,1,1,1,1]
    random.shuffle(probAnsSide)
    
    sheet.cell(row = 0,column = 0).value = 'type'
    
    if type == 'p' or type == 'ps' or type == 'pm':
        sheet.cell(row = 0,column=1).value = 'letter1'
        sheet.cell(row = 0,column=2).value = 'letter2'
        sheet.cell(row = 0,column=3).value = 'letter3'
        sheet.cell(row = 0,column=4).value = 'letter4'
        sheet.cell(row = 0,column=5).value = 'letter5'
        sheet.cell(row = 0,column=6).value = 'letter6'
        sheet.cell(row = 0,column=7).value = 'letter7'        
        sheet.cell(row = 0,column=8).value = 'letter8'
        sheet.cell(row = 0,column=9).value = 'letter9'
        sheet.cell(row = 0,column=10).value = 'loc1x'
        sheet.cell(row = 0,column=11).value = 'loc1y'
        sheet.cell(row = 0,column=12).value = 'loc2x'
        sheet.cell(row = 0,column=13).value = 'loc2y'
        sheet.cell(row = 0,column=14).value = 'loc3x'
        sheet.cell(row = 0,column=15).value = 'loc3y'
        sheet.cell(row = 0,column=16).value = 'loc4x'
        sheet.cell(row = 0,column=17).value = 'loc4y'
        sheet.cell(row = 0,column=18).value = 'loc5x'
        sheet.cell(row = 0,column=19).value = 'loc5y'
        sheet.cell(row = 0,column=20).value = 'loc6x'
        sheet.cell(row = 0,column=21).value = 'loc6y'
        sheet.cell(row = 0,column=22).value = 'loc7x'
        sheet.cell(row = 0,column=23).value = 'loc7y'
        sheet.cell(row = 0,column=24).value = 'loc8x'
        sheet.cell(row = 0,column=25).value = 'loc8y'
        sheet.cell(row = 0,column=26).value = 'loc9x'
        sheet.cell(row = 0,column=27).value = 'loc9y'   
        sheet.cell(row = 0,column=28).value = 'span'
        sheet.cell(row = 0,column=29).value = 'spanLevel'
        sheet.cell(row = 0,column=30).value = 'probeLet1'
        sheet.cell(row = 0,column=31).value = 'probeLet2'
        sheet.cell(row = 0,column=32).value = 'taskAns'
    elif type == 'v' or type == 'vs' or type == 'vm':
        sheet.cell(row = 0,column=1).value = 'loc1'
        sheet.cell(row = 0,column=2).value = 'loc2'
        sheet.cell(row = 0,column=3).value = 'loc3'
        sheet.cell(row = 0,column=4).value = 'loc4'
        sheet.cell(row = 0,column=5).value = 'loc5'
        sheet.cell(row = 0,column=6).value = 'loc6'
        sheet.cell(row = 0,column=7).value = 'loc7'
        sheet.cell(row = 0,column=8).value = 'loc8'
        sheet.cell(row = 0,column=9).value = 'loc9'    
        sheet.cell(row = 0,column=10).value = 'span'
        sheet.cell(row = 0,column=11).value = 'spanLevel'
        sheet.cell(row = 0,column=12).value = 'taskAns'
        sheet.cell(row = 0,column=13).value = 'probeLoc'
        sheet.cell(row = 0,column=14).value = 'probeIndx'
        sheet.cell(row = 0,column=15).value = 'dir'        
    if type == 'vs' or type == 'vm':
        sheet.cell(row = 0,column=16).value = 'operand1'
        sheet.cell(row = 0,column=17).value = 'operand2'    
        sheet.cell(row = 0,column=18).value = 'probRes'
        sheet.cell(row = 0,column=19).value = 'respAlt'
        sheet.cell(row = 0,column=20).value = 'probAnsSide'
        sheet.cell(row = 0,column=21).value = 'subDist'
        sheet.cell(row = 0,column=22).value = 'tableError'
        sheet.cell(row = 0,column=23).value = 'carry'
    if type == 'pm' or type == 'ps':
        sheet.cell(row = 0,column=33).value = 'operand1'
        sheet.cell(row = 0,column=34).value = 'operand2'    
        sheet.cell(row = 0,column=35).value = 'probRes'
        sheet.cell(row = 0,column=36).value = 'respAlt'
        sheet.cell(row = 0,column=37).value = 'probAnsSide'
        sheet.cell(row = 0,column=38).value = 'subDist'
        sheet.cell(row = 0,column=39).value = 'tableError'
        sheet.cell(row = 0,column=40).value = 'carry'

    locsX = []
    locsY = []
    line = 1
    for j in range(0,2):
        span = spans[j]
        if type == 'p':
            for i in range(0,reps):
                col = 0
                sheet.cell(row = line,column=0).value = 'p'
                sheet.cell(row = line,column=28).value = span #write the span
                sheet.cell(row = line,column=29).value = j+1 #write the span level
                lets = wmLets(span)
                locs = phonoLocs(span)        
                usedLets = [] #the array that keeps the letters used in the current trial
                for col in range(0,span):
                    sheet.cell(row = line,column=col+1).value =  listOfConsonants[int(lets[col])] #write the letter to be used in the current trial
                    usedLets.append(listOfConsonants[int(lets[col])])
                for col in range(10,10+span*2):
                    if col%2 == 1:
                        sheet.cell(row = line,column=col).value =  0 #write the y location for the current letter
                    else:
                        sheet.cell(row = line,column=col).value =  locs[(col-10)/2] #write the x location for the current letter
                
                if i < reps/2:
                    sheet.cell(row = line,column=30).value = -1 #no prob let for correct sequence
                    sheet.cell(row = line,column=31).value = -1 #no prob let for correct sequence
                    sheet.cell(row = line,column=32).value = corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                else:
                    tmp = numpy.zeros(span)
                    for i in range(0,span):
                        tmp[i] = i
                    #print 'tmp',tmp
                    let1 = choice(tmp)
                    let2 = choice(tmp)
                    while let1== let2:
                        let1 = choice(tmp) 
                        #print 'let1',let1,'let2',let2
                    sheet.cell(row = line,column=30).value = int(let1) #pick the location of the letter to be swapped
                    sheet.cell(row = line,column=31).value = int(let2) #pick the location of the letter to be swapped
                    sheet.cell(row = line,column=32).value = 1-corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                line = line + 1
            
        elif type == 'v':
            for i in range(0,reps):
                sheet.cell(row = line,column=0).value = 'v'
                sheet.cell(row = line,column=10).value = span#span
                sheet.cell(row = line,column=11).value = j+1#span level
                col = 0
                lets = wmLets(span)
                locsUsed = visualLocs(span)
                if stepDirV[i%20] != 0:
                    indx,probLoc,dir = findProb(locsUsed,stepDirV[i%20],span)
                else:
                    indx = -1
                    probLoc = -1
                    dir = 0
                for col in range(0,span):
                        sheet.cell(row = line,column=col+1).value =  locsUsed[col] #write the x location for the current letter               
                sheet.cell(row = line,column=13).value = probLoc
                sheet.cell(row = line,column=14).value = indx
                sheet.cell(row = line,column=15).value = dir
                if stepDirV[i%20] == 0:
                    sheet.cell(row = line,column=12).value = corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                else:
                    sheet.cell(row = line,column=12).value = 1-corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                line = line + 1
                
        elif type == 'ps':
            for i in range(0,reps):
                col = 0
                sheet.cell(row = line,column=0).value = 'ps'
                sheet.cell(row = line,column=28).value = span #write the span
                sheet.cell(row = line,column=29).value = j+1 #write the span level
                sheet.cell(row = line,column=32).value = int(i%span) #the location of the probe letter within the letters list
                lets = wmLets(span)
                locs = phonoLocs(span)        
                usedLets = [] #the array that keeps the letters used in the current trial
                for col in range(0,span):
                    sheet.cell(row = line,column=col+1).value =  listOfConsonants[int(lets[col])] #write the letter to be used in the current trial
                    usedLets.append(listOfConsonants[int(lets[col])])
                for col in range(10,10+span*2):
                    if col%2 == 1:
                        sheet.cell(row = line,column=col).value =  0 #write the y location for the current letter
                    else:
                        sheet.cell(row = line,column=col).value =  locs[(col-10)/2] #write the x location for the current letter
                if i < reps/2:
                    sheet.cell(row = line,column=30).value = -1 #for the correct test letters, make sure that each of the shown letters are used equally as the test
                    sheet.cell(row = line,column=31).value = -1 #for the correct test letters, make sure that each of the shown letters are used equally as the test                    
                    sheet.cell(row = line,column=32).value = corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                else:
                    tmp = numpy.zeros(span)
                    for i in range(0,span):
                        tmp[i] = i
                    let1 = choice(tmp)
                    let2 = choice(tmp)
                    while let1== let2:
                        #print 'BURA'
                        let1 = choice(tmp)  
                    sheet.cell(row = line,column=30).value = int(let1) #pick the location of the letter to be swapped
                    sheet.cell(row = line,column=31).value = int(let2) #pick the location of the letter to be swapped
                    sheet.cell(row = line,column=32).value = 1-corrAnsSide #test letter is not one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->right(0), even-->left(1)
                sheet.cell(row = line,column=33).value = calcSheet.cell(row=line,column =1).value#operand1[line]
                sheet.cell(row = line,column=34).value = calcSheet.cell(row=line,column =2).value#operand2[line]
                sheet.cell(row = line,column=35).value = calcSheet.cell(row=line,column =3).value#probRes[line]
                sheet.cell(row = line,column=36).value = calcSheet.cell(row=line,column =4).value#respAlt[line]
                sheet.cell(row = line,column=37).value = probAnsSide[line%28-1]
                sheet.cell(row = line,column=38).value = calcSheet.cell(row=line,column =5).value#subDist[line]
                sheet.cell(row = line,column=39).value = calcSheet.cell(row=line,column =6).value#'-1' #empty cell for tableError
                sheet.cell(row = line,column=40).value = calcSheet.cell(row=line,column =7).value#carry
                line = line + 1

        elif type == 'vs':
            for i in range(0,reps):
                sheet.cell(row = line,column=0).value = 'vs'
                sheet.cell(row = line,column=10).value = span#span
                sheet.cell(row = line,column=11).value = j+1#span level
                col = 0
                lets = wmLets(span)
                locsUsed = visualLocs(span)        
                if stepDirV[i%20] != 0:
                    indx,probLoc,dir = findProb(locsUsed,stepDirV[i%20],span)
                else:
                    indx = -1
                    probLoc = -1
                    dir = 0
                for col in range(0,span):
                        sheet.cell(row = line,column=col+1).value =  locsUsed[col] #write the x location for the current letter
                sheet.cell(row = line,column=13).value = probLoc
                sheet.cell(row = line,column=14).value = indx
                sheet.cell(row = line,column=15).value = dir
                if stepDirVC[i%28] == 0:
                    sheet.cell(row = line,column=12).value = corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                else:
                    sheet.cell(row = line,column=12).value = 1-corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                sheet.cell(row = line,column=16).value = calcSheet.cell(row=line,column =1).value#operand1[line]
                sheet.cell(row = line,column=17).value = calcSheet.cell(row=line,column =2).value#operand2[line]
                sheet.cell(row = line,column=18).value = calcSheet.cell(row=line,column =3).value#probRes[line]
                sheet.cell(row = line,column=19).value = calcSheet.cell(row=line,column =4).value#respAlt[line]
                sheet.cell(row = line,column=20).value = probAnsSide[line%28-1]
                sheet.cell(row = line,column=21).value = calcSheet.cell(row=line,column =5).value#subDist[line]
                sheet.cell(row = line,column=22).value = calcSheet.cell(row=line,column =6).value#tableError[line] #empty space for tableError
                sheet.cell(row = line,column=23).value = calcSheet.cell(row=line,column =7).value#carry
                line = line + 1
        elif type == 'pm':
            for i in range(0,reps):
                col = 0
                sheet.cell(row = line,column=0).value = 'pm'
                sheet.cell(row = line,column=28).value = span #write the span
                sheet.cell(row = line,column=29).value = j+1 #write the span level
                sheet.cell(row = line,column=32).value =  int(i%span) #the location of the probe letter within the letters list
                lets = wmLets(span)
                locs = phonoLocs(span)        
                usedLets = [] #the array that keeps the letters used in the current trial
                for col in range(0,span):
                    sheet.cell(row = line,column=col+1).value =  listOfConsonants[int(lets[col])] #write the letter to be used in the current trial
                    usedLets.append(listOfConsonants[int(lets[col])])
                for col in range(10,10+span*2):
                    if col%2 == 1:
                        sheet.cell(row = line,column=col).value =  0 #write the y location for the current letter
                    else:
                        sheet.cell(row = line,column=col).value =  locs[(col-10)/2] #write the x location for the current letter
                if i < reps/2:
                    sheet.cell(row = line,column=30).value = -1 #for the correct test letters, make sure that each of the shown letters are used equally as the test
                    sheet.cell(row = line,column=31).value = -1 #for the correct test letters, make sure that each of the shown letters are used equally as the test                    
                    sheet.cell(row = line,column=32).value = corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                else:
                    tmp = numpy.zeros(span)
                    for i in range(0,span):
                        tmp[i] = i
                    let1 = choice(tmp)
                    let2 = choice(tmp)
                    while let1== let2:
                        #print 'BURA1'
                        let1 = choice(tmp)  
                    sheet.cell(row = line,column=30).value = int(let1) #pick the location of the letter to be swapped
                    sheet.cell(row = line,column=31).value = int(let2) #pick the location of the letter to be swapped
                    sheet.cell(row = line,column=32).value = 1-corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                sheet.cell(row = line,column=33).value = calcSheet.cell(row=line,column =1).value#operand1[line]
                sheet.cell(row = line,column=34).value = calcSheet.cell(row=line,column =2).value#operand2[line]
                sheet.cell(row = line,column=35).value = calcSheet.cell(row=line,column =3).value#probRes[line]
                sheet.cell(row = line,column=36).value = calcSheet.cell(row=line,column =4).value#respAlt[line]
                sheet.cell(row = line,column=37).value = probAnsSide[line%28-1]
                sheet.cell(row = line,column=38).value = calcSheet.cell(row=line,column =5).value#subDist[line]
                sheet.cell(row = line,column=39).value = calcSheet.cell(row=line,column =6).value#'-1' #empty cell for tableError
                sheet.cell(row = line,column=40).value = calcSheet.cell(row=line,column =7).value#carry
                line = line + 1
                
        elif type == 'vm':
            for i in range(0,reps):
                sheet.cell(row = line,column=0).value = 'vm'
                sheet.cell(row = line,column=10).value =span#span
                sheet.cell(row = line,column=11).value = j+1#span level
                col = 0
                lets = wmLets(span)
                locsUsed = visualLocs(span)
                if stepDirV[i%20] != 0:
                    indx,probLoc,dir = findProb(locsUsed,stepDirV[i%20],span)
                else:
                    indx = -1
                    probLoc = -1
                    dir = 0
                for col in range(0,span):
                    sheet.cell(row = line,column=col+1).value =  locsUsed[col] #write the x location for the current letter
                sheet.cell(row = line,column=13).value = probLoc
                sheet.cell(row = line,column=14).value = indx
                sheet.cell(row = line,column=15).value = dir
                if stepDirVC[i%28] == 0:
                    sheet.cell(row = line,column=12).value = corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                else:
                    sheet.cell(row = line,column=12).value = 1-corrAnsSide #test letter is one of the previously shown letters (left or right button press depending on the parity of the participant number),odd-->left, even-->right
                sheet.cell(row = line,column=16).value = calcSheet.cell(row=line,column =1).value#operand1[line]
                sheet.cell(row = line,column=17).value = calcSheet.cell(row=line,column =2).value#operand2[line]
                sheet.cell(row = line,column=18).value = calcSheet.cell(row=line,column =3).value#probRes[line]
                sheet.cell(row = line,column=19).value = calcSheet.cell(row=line,column =4).value#respAlt[line]
                sheet.cell(row = line,column=20).value = probAnsSide[line%28-1]
                sheet.cell(row = line,column=21).value = calcSheet.cell(row=line,column =5).value#subDist[line]
                sheet.cell(row = line,column=22).value = calcSheet.cell(row=line,column =6).value#tableError[line] #empty space for tableError
                sheet.cell(row = line,column=23).value = calcSheet.cell(row=line,column =7).value#carry
                line = line + 1

    #To save as a different file:
    #if os.path.exists(file_name):
        #os.remove(file_name)
    wb.save('C:\\Users\\user\\Desktop\\exp\\mydata\\subject_'+str(participant)+'_session_'+type+'.xlsx')#change name

#run the create stimuli code and obtain the excel file with stimuli information
def run_create_stimuli(participant,group,calcFile,type,reps):
#    if not os.path.isdir('v_staircase'):
#        os.makedirs('v_staircase')  # if this fails (e.g. permissions) we will get error
#    vFileName = 'v_staircase' + os.path.sep + 'subject_' + str(participant) + '.psydat'
#    allIntensities, allResponses = [],[]
#    thisDat = misc.fromFile(vFileName)
#  
#    #get the data from all the files
#    allIntensities, allResponses = [],[]
#    combinedInten = [1,2,3,4,5,6,7]
#    combinedCorrResp = [0,0,0,0,0,0,0]
#    combinedIncorrResp = [0,0,0,0,0,0,0]
#    combinedResp = [0,0,0,0,0,0,0]
#    finalResp = []
#    finalInten = []
#    for i in range(0,len(thisDat.intensities)):
#        if thisDat.intensities[i]> 7:
#            thisDat.intensities[i] = 7
#    for i in range(0,len(thisDat.intensities)):
#        if thisDat.data[i] == 1:
#            combinedCorrResp[thisDat.intensities[i]-1] += 1
#        else:
#            combinedIncorrResp[thisDat.intensities[i]-1] += 1
#    intsToBeDeleted = []
#    for i in range(1,8):
#        tmp1 = combinedCorrResp[i-1]+combinedIncorrResp[i-1]
#        print 'tmp1',tmp1
#        if tmp1 != 0:
#            combinedResp[i-1] = combinedCorrResp[i-1]/tmp1
#        else:
#            intsToBeDeleted.append(i)
#
#    for i in range(0,7):
#        add = 1
#        for j in range(0,len(intsToBeDeleted)):
#            if combinedInten[i] == intsToBeDeleted[j]:
#                add = 0
#        if add == 1:
#            finalInten.append(combinedInten[i])
#            finalResp.append(combinedResp[i])
#        
#
#    #get combined data
#    print 'allintentisites ', allIntensities, 'allResponses ',allResponses
#    #combinedInten, combinedResp, combinedN = \
#    #             data.functionFromStaircase(allIntensities, allResponses,6)
#
#    #fit curve - in this case using a Weibull function
#    guess= [7,2]
#    #fit curve - in this case using a Weibull function
#    #fit = data.FitFunction('weibullYN',combinedInten, combinedResp, \
#    #guess = guess)
#    print 'finalInten',finalInten, 'finalResp',finalResp#,'combinedN',combinedN
#    fit = data.FitFunction('weibullYN',finalInten, finalResp, \
#    guess = guess)
#    #fit = data.FitWeibull(finalInten,finalResp)#,guess=guess)
#    #guess = guess)
#    #guess=[0.2, 0.5])
#
#
#    vSpan1 = fit.inverse(0.99)
#    vSpan2 = fit.inverse(0.80)
#    print 'real thresholds', vSpan1,vSpan2
#
#    if vSpan1 > 6: vSpan1 = 6
#    elif vSpan1 <=0: vSpan1 = 3
#    if vSpan2 >7: vSpan2 = 7
#    elif vSpan2 <=0: vSpan2 = 5
#    
#    vSpan1 = int(vSpan1)
#    vSpan2 = int(vSpan2)
#    
#    
#    if vSpan1 > 6: vSpan1 = 6
#    elif vSpan1 <=0: vSpan1 = 3
#    if vSpan2 >7: vSpan2 = 7
#    elif vSpan2 <=0: vSpan2 = 5
#
#    if vSpan1 == vSpan2: vSpan1 -= 1
#    print 'thresholds v: ',vSpan1,vSpan2
#    


    if not os.path.isdir('p_staircase'):
        os.makedirs('p_staircase')  # if this fails (e.g. permissions) we will get error
    pFileName = 'p_staircase' + os.path.sep + 'subject_' + str(participant) + '.psydat'

    #get the data from all the files
    allIntensities, allResponses = [],[]
    combinedInten = [1,2,3,4,5,6,7,8,9]
    combinedCorrResp = [0,0,0,0,0,0,0,0,0]
    combinedIncorrResp = [0,0,0,0,0,0,0,0,0]
    combinedResp = [0,0,0,0,0,0,0,0,0]
    finalResp = []
    finalInten = []
    thisDat = misc.fromFile(pFileName)
    allIntensities.append(thisDat.intensities)
    allResponses.append(thisDat.data)
    print 'intensities: ',thisDat.intensities
    print len(thisDat.intensities)
    for i in range(0,len(thisDat.intensities)):
        if thisDat.intensities[i]> 9:
            thisDat.intensities[i] = 9
    for i in range(0,len(thisDat.intensities)):
        if thisDat.data[i] == 1:
            combinedCorrResp[thisDat.intensities[i]-1] += 1
        else:
            combinedIncorrResp[thisDat.intensities[i]-1] += 1
    intsToBeDeleted = []
    for i in range(1,10):
        tmp1 = combinedCorrResp[i-1]+combinedIncorrResp[i-1]
        print 'tmp1',tmp1
        if tmp1 != 0:
            combinedResp[i-1] = combinedCorrResp[i-1]/tmp1
        else:
            intsToBeDeleted.append(i)

    for i in range(0,9):
        add = 1
        for j in range(0,len(intsToBeDeleted)):
            if combinedInten[i] == intsToBeDeleted[j]:
                add = 0
        if add == 1:
            finalInten.append(combinedInten[i])
            finalResp.append(combinedResp[i])
       


    #get combined data
    print 'allintentisites ', allIntensities, 'allResponses ',allResponses


    #fit curve - in this case using a Weibull function
    guess= [9,2]

    print 'finalInten',finalInten, 'finalResp',finalResp#,'combinedN',combinedN
    fit = data.FitFunction('weibullYN',finalInten, finalResp, \
    guess = guess)
    #fit = data.FitWeibull(finalInten,finalResp)#,guess=guess)
    #guess = guess)
    #guess=[0.2, 0.5])
    pSpan1 = fit.inverse(0.99)
    pSpan2 = fit.inverse(0.80)
    print 'real thresholds',pSpan1,pSpan2
    
    if pSpan1 > 8: pSpan1 = 8
    elif pSpan1 <= 0: pSpan1 = 5
    if pSpan2 > 9: pSpan2 = 9
    elif pSpan2 <=0: pSpan2 = 7
    
    

    pSpan1 = int(pSpan1)
    pSpan2 = int(pSpan2)
    
    if pSpan1 > 8: pSpan1 = 8
    elif pSpan1 <= 0: pSpan1 = 5
    if pSpan2 > 9: pSpan2 = 9
    elif pSpan2 <=0: pSpan2 = 7
    elif pSpan1 <=1: pSpan1 = 2
    elif pSpan2<=1: pSpan2 = 2
    
    if pSpan1 == pSpan2: pSpan1 -= 1
    print 'thresholds: ',pSpan1, pSpan2

    spans = []
    #use pspans if the type is p, ps or pm
    if type == 'p' or type == 'ps' or type == 'pm':
        spans = [pSpan1,pSpan2]
    else:
        spans = [vSpan1,vSpan2]
    #print spans
    create_stimuli(participant,group,calcFile,type,reps,spans)
    #after the experiment save the data
    print '\n'